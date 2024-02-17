#!/usr/bin/env python3

import os
import warnings
import random
import re
import tempfile
import checklib
import requests
import openpyxl
import json
import sys
from checklib import *
from websockets.sync.client import connect
from websockets.exceptions import WebSocketException
from checklib import status

import cell_lib

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

class Checker(BaseChecker):
    vulns: int = 1
    timeout: int = 20
    uses_attack_data: bool = True

    def __init__(self, *args, **kwargs):
        super(Checker, self).__init__(*args, **kwargs)
        self.cm = cell_lib.CheckMachine(self)
        self.uuid_regexp = re.compile(r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$')
        self.token_regexp = re.compile(r'^[0-9A-Za-z]{1,40}$')

    def decode_json(
        self, s: str, public: str, status=checklib.status.Status.MUMBLE):
        json_errors = (
            UnicodeDecodeError,
            json.decoder.JSONDecodeError,
            ValueError,
        )
        try:
            data = json.loads(s)
        except json_errors:
            self.cquit(status, public, f"Invalid JSON while decoding: {s[:100]}")
        else:
            return data
    
    def _create_excel_file(self, sheet_name:str, cell_value:str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = cell_value
        ws.title = sheet_name
            
        fp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        fp.close()
        wb.save(fp.name)
        return fp
        

    def check(self):
        sess = get_initialized_session()
        username, password = rnd_username(), rnd_password()
        self.cm.register(sess, username, password)
        self.cm.login(sess, username, password)

        sheet_name = rnd_string(31)
        cell_value = rnd_string(31)
        upload_file = random.randint(0, 1) == 1

        if upload_file == 1:
            fp = self._create_excel_file(sheet_name, cell_value)
            sheet = self.cm.upload_sheet(sess, sheet_name + ".xlsx", fp.name)
            os.unlink(fp.name)
        else:
            sheet = self.cm.create_sheet(sess, sheet_name)
        sid = sheet.get('sid')
        readToken = sheet.get('readToken')
        modifyToken = sheet.get('modifyToken')

        self.assert_eq(bool(self.uuid_regexp.fullmatch(sid)), True, 'Invalid sheet id format')
        self.assert_eq(bool(self.token_regexp.fullmatch(readToken)), True, 'Invalid read token format')
        self.assert_eq(bool(self.token_regexp.fullmatch(modifyToken)), True, 'Invalid modify token format')

        
        self.assert_eq(sheet.get('title'), sheet_name, 'Invalid sheet title returned')

        get_sheet_resp = self.cm.get_sheet(sess, sid, readToken)
        self.assert_eq(get_sheet_resp.get('title'), sheet_name, 'Invalid sheet title returned')
        if upload_file == 1:
            values = [x.get('val', '') for x in get_sheet_resp.get('cells', [])]
            self.assert_in(cell_value, values, 'Cell value not found while fetching sheet')

        with connect(self.cm.ws_url) as ws_writer, connect(self.cm.ws_url) as ws_reader:
            # Connect writer.
            self.cm.connect(ws_writer, 1)
            
            # Connect reader.
            self.cm.connect(ws_reader, 1)
            
            # Subscribe to sheet.
            sub_resp = self.cm.subscribe(ws_reader, 2, sid, readToken)
            self.assert_in('subscribe', sub_resp.keys(), 'Invalid subscribe response')
        
            new_cell_value = rnd_string(30)
            # Modify cell.
            # Try to modify cell twice to solve the reliability issue.
            for _attempt in range(0, 2):
                try:
                    _modify_result = self.cm.modify_cell(ws_writer, 2, sid, 'B2', new_cell_value, modifyToken, timeout=self.timeout // 3)
                    break
                except TimeoutError:
                    continue

            self.assert_neq(_modify_result, None, 'Failed to modify cell: no response received')
            self.assert_in('rpc', _modify_result.keys(), 'Invalid modify_cell response')
            mb_sheet = _modify_result.get('rpc', {}).get('data', {}).get('sheet', {})
            self.assert_eq(mb_sheet.get('title', None), sheet_name, 'Invalid sheet title in modify_cell response')
            cell_values = [x.get('val', '') for x in mb_sheet.get('cells', [])]
            self.assert_in(new_cell_value, cell_values, 'New cell value not found in modify_cell response')

            # Get update from reader.
            # Try to get update twice to solve the reliability issue.
            for _attempt in range(0, 2):
                expected_update = ws_reader.recv(timeout=self.timeout // 3)
                update_data = self.decode_json(expected_update, 'Invalid JSON received from sheet update')
                if update_data:
                    break

            self.assert_eq(type(update_data), dict, 'Invalid JSON received from sheet update')
            update_data = update_data.get('push', {}).get('pub', {}).get('data', {})
            self.assert_eq(update_data.get('title', ''), sheet_name, 'Invalid sheet title returned from sheet update')
            cell_values = [x.get('val', '') for x in update_data.get('cells', [])]
            self.assert_in(new_cell_value, cell_values, 'New cell value not found in sheet update response')


        sheets = self.cm.user_sheets(sess)
        self.assert_eq(len(sheets), 1, 'Invalid sheets count returned')
        self.assert_eq(sheets[0].get('title'), sheet_name, 'Failed to get user sheets: invalid title returned')
        self.assert_eq(sheets[0].get('sid'), sid, 'Failed to get user sheets: invalid id returned')
        self.assert_eq(sheets[0].get('readToken'), readToken, 'Failed to get user sheets: invalid token returned')
        self.assert_eq(sheets[0].get('modifyToken'), modifyToken, 'Failed to get user sheets: invalid token returned')

        self.cquit(Status.OK)
    
    def put(self, flag_id: str, flag: str, vuln: str):
        sess = get_initialized_session()
        username, password = rnd_username(), rnd_password()
        self.cm.register(sess, username, password)
        self.cm.login(sess, username, password)

        upload_file = random.randint(0, 1) == 1
        if upload_file == 1:
            sheet = self.cm.create_sheet(sess, flag)
        else:
            fp = self._create_excel_file(flag, "Flag is not: " + rnd_string(31))
            sheet = self.cm.upload_sheet(sess, rnd_string(31) + ".xlsx", fp.name)
            os.unlink(fp.name)

        sid = sheet.get('sid')
        readToken = sheet.get('readToken')
        self.assert_eq(bool(self.uuid_regexp.fullmatch(sid)), True, 'Invalid sheet id format')
        self.assert_eq(bool(self.token_regexp.fullmatch(readToken)), True, 'Invalid read token format')

        self.cquit(Status.OK, sid, f"{username}:{password}:{sid}:{readToken}")



    def get(self, flag_id: str, flag: str, vuln: str):
        username, password, sid, token = flag_id.split(':')
        sess = get_initialized_session()
        
        get_sheet_resp = self.cm.get_sheet(sess, sid, token)
        self.assert_eq(get_sheet_resp.get('title'), flag, 'Invalid sheet title returned', status=Status.CORRUPT)

        self.cm.login(sess, username, password, status=Status.CORRUPT)
        user_sheets = self.cm.user_sheets(sess, status=Status.CORRUPT)

        titles = [x.get('title') for x in user_sheets]
        sids = [x.get('sid') for x in user_sheets]
        tokens = [x.get('readToken') for x in user_sheets]

        self.assert_in(sid, sids, 'Sheet not found in user sheets', status=Status.CORRUPT)
        self.assert_in(flag, titles, 'Sheet not found in user sheets', status=Status.CORRUPT)
        self.assert_in(token, tokens, 'Sheet not found in user sheets', status=Status.CORRUPT)

        self.cquit(Status.OK)
        
    
    def action(self, action, *args, **kwargs):
        try:
            super(Checker, self).action(action, *args, **kwargs)
        except requests.exceptions.ConnectionError:
            self.cquit(Status.DOWN, 'Connection error', 'Got requests connection error')
        except WebSocketException:
            self.cquit(Status.DOWN, 'Connection error', 'Got ws connection error')
    


if __name__ == '__main__':
    c = Checker(sys.argv[2])

    try:
        c.action(sys.argv[1], *sys.argv[3:])
    except c.get_check_finished_exception():
        cquit(Status(c.status), c.public, c.private)
